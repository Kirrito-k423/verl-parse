#!/usr/bin/env python3
"""Parse two VERL logs and export per-step metrics with a compare sheet."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ANSI_ESCAPE_RE = re.compile(r"\x1b\[[0-?]*[ -/]*[@-~]")
METRIC_RE = re.compile(
    r"(?P<key>[A-Za-z0-9_./@-]+):\s*"
    r"(?P<value>[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?)"
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
COMPARE_FILL = PatternFill(fill_type="solid", fgColor="EAF4E4")
MISSING_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
GRAY_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
PCT_FORMAT = "0.00%"

SUMMARY_METRICS: List[tuple[str, Sequence[str]]] = [
    ("critic/rewards/mean", ("critic/rewards/mean",)),
    ("response_length/mean", ("response_length/mean",)),
    (
        "training/rollout_probs_diff_mean",
        ("training/rollout_probs_diff_mean",),
    ),
    ("actor/grad_norm", ("actor/grad_norm",)),
    ("actor/kl_loss", ("actor/kl_loss",)),
    ("prompt_length/mean", ("prompt_length/mean",)),
    (
        "val-aux/hiyouga/geometry3k/reward/mean@1",
        ("val-aux/hiyouga/geometry3k/reward/mean@1",),
    ),
    ("actor/lr", ("actor/lr",)),
    ("timing_s/gen", ("timing_s/gen",)),
    ("timing_s/logp", ("timing_s/log_prob", "timing_s/logp", "timing_s/ref")),
    ("timing_s/old_log_prob", ("timing_s/old_log_prob",)),
    ("timing_s/update_actor", ("timing_s/update_actor",)),
    ("timing_s/step", ("timing_s/step",)),
    ("perf/throughput", ("perf/throughput",)),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Parse two VERL log files and export per-step metrics into Excel "
            "with a compare sheet."
        )
    )
    parser.add_argument("log1", type=Path, help="First VERL log file.")
    parser.add_argument("log2", type=Path, help="Second VERL log file.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("verl_log_compare.xlsx"),
        help="Output Excel file path. Default: verl_log_compare.xlsx",
    )
    parser.add_argument(
        "--label1",
        type=str,
        default=None,
        help="Display label for the first log sheet/compare column.",
    )
    parser.add_argument(
        "--label2",
        type=str,
        default=None,
        help="Display label for the second log sheet/compare column.",
    )
    return parser.parse_args()


def clean_line(raw_line: str) -> str:
    return ANSI_ESCAPE_RE.sub("", raw_line).replace("\r", " ").strip()


def to_number(raw_value: str) -> float | int:
    if re.fullmatch(r"[-+]?\d+", raw_value):
        return int(raw_value)
    return float(raw_value)


def parse_log(log_path: Path) -> Dict[int, Dict[str, float | int]]:
    if not log_path.is_file():
        raise FileNotFoundError(f"Log file not found: {log_path}")

    step_to_metrics: Dict[int, Dict[str, float | int]] = {}

    with log_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            line = clean_line(raw_line)
            if "step:" not in line:
                continue

            metrics: Dict[str, float | int] = {}
            for match in METRIC_RE.finditer(line):
                metrics[match.group("key")] = to_number(match.group("value"))

            step_value = metrics.get("step")
            if step_value is None:
                continue
            if len(metrics) <= 1:
                continue

            step = int(step_value)
            if step in step_to_metrics:
                print(
                    (
                        f"Warning: duplicate step {step} found in {log_path} at "
                        f"line {line_number}; keeping the latest record."
                    ),
                    file=sys.stderr,
                )
            step_to_metrics[step] = metrics

    if not step_to_metrics:
        raise ValueError(f"No step metrics were found in log file: {log_path}")

    return dict(sorted(step_to_metrics.items(), key=lambda item: item[0]))


def collect_metric_keys(step_map: Dict[int, Dict[str, float | int]]) -> List[str]:
    metric_keys = set()
    for metrics in step_map.values():
        metric_keys.update(metrics.keys())
    metric_keys.discard("step")
    return sorted(metric_keys)


def ensure_distinct_labels(label1: str, label2: str) -> tuple[str, str]:
    if label1 != label2:
        return label1, label2
    return f"{label1}_1", f"{label2}_2"


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[:\\/?*[\]]", "_", title).strip() or "sheet"
    return cleaned[:31]


def apply_header_style(cell, fill: PatternFill) -> None:
    cell.font = Font(bold=True)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_fit_columns(ws) -> None:
    max_widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            width = min(max(len(value) + 2, 10), 40)
            max_widths[cell.column] = max(max_widths.get(cell.column, 0), width)

    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def resolve_metric(
    metrics: Dict[str, float | int],
    aliases: Sequence[str],
) -> tuple[Optional[str], Optional[float | int]]:
    for alias in aliases:
        if alias in metrics:
            return alias, metrics[alias]
    return None, None


def compare_values(
    value1: Optional[float | int],
    value2: Optional[float | int],
) -> tuple[Optional[float], Optional[float], str]:
    diff: Optional[float] = None
    pct_diff: Optional[float] = None
    status = "ok"

    if value1 is None and value2 is None:
        return diff, pct_diff, "missing_in_both"
    if value1 is None:
        return diff, pct_diff, "missing_in_log1"
    if value2 is None:
        return diff, pct_diff, "missing_in_log2"

    diff = float(value2) - float(value1)
    if float(value1) == 0.0:
        if float(value2) == 0.0:
            pct_diff = 0.0
        else:
            status = "baseline_zero"
    else:
        pct_diff = diff / float(value1)

    return diff, pct_diff, status


def fill_pct_cell(cell, pct_diff: Optional[float], status: str) -> None:
    if status not in {"ok", "baseline_zero"} or pct_diff is None:
        cell.fill = GRAY_FILL
        return

    abs_pct = abs(pct_diff)
    if abs_pct <= 0.01:
        cell.fill = GREEN_FILL
    elif abs_pct <= 0.05:
        cell.fill = YELLOW_FILL
    else:
        cell.fill = RED_FILL


def write_log_sheet(
    ws,
    *,
    step_map: Dict[int, Dict[str, float | int]],
    metric_keys: List[str],
    source_path: Path,
) -> None:
    ws.append(["source_file", str(source_path.resolve())])
    ws.append(["step", *metric_keys])

    for cell in ws[2]:
        apply_header_style(cell, HEADER_FILL)

    for step, metrics in step_map.items():
        row = [step]
        row.extend(metrics.get(metric) for metric in metric_keys)
        ws.append(row)

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = ws.dimensions
    auto_fit_columns(ws)


def build_compare_rows(
    log1_steps: Dict[int, Dict[str, float | int]],
    log2_steps: Dict[int, Dict[str, float | int]],
    metric_keys: Iterable[str],
) -> List[List[object]]:
    rows: List[List[object]] = []
    all_steps = sorted(set(log1_steps) | set(log2_steps))

    for step in all_steps:
        metrics1 = log1_steps.get(step, {})
        metrics2 = log2_steps.get(step, {})
        for metric in metric_keys:
            value1 = metrics1.get(metric)
            value2 = metrics2.get(metric)
            if value1 is None and value2 is None:
                continue

            diff, pct_diff, status = compare_values(value1, value2)

            rows.append([step, metric, value1, value2, diff, pct_diff, status])

    return rows


def build_summary_rows(
    log1_steps: Dict[int, Dict[str, float | int]],
    log2_steps: Dict[int, Dict[str, float | int]],
) -> List[List[object]]:
    rows: List[List[object]] = []
    all_steps = sorted(set(log1_steps) | set(log2_steps))

    for step in all_steps:
        metrics1 = log1_steps.get(step, {})
        metrics2 = log2_steps.get(step, {})
        for display_name, aliases in SUMMARY_METRICS:
            metric1_name, value1 = resolve_metric(metrics1, aliases)
            metric2_name, value2 = resolve_metric(metrics2, aliases)
            diff, pct_diff, status = compare_values(value1, value2)
            rows.append(
                [
                    step,
                    display_name,
                    metric1_name,
                    metric2_name,
                    value1,
                    value2,
                    diff,
                    pct_diff,
                    status,
                ]
            )

    return rows


def write_summary_sheet(
    ws,
    *,
    log1_path: Path,
    log2_path: Path,
    label1: str,
    label2: str,
    rows: List[List[object]],
) -> None:
    ws.append(["baseline_log", str(log1_path.resolve())])
    ws.append(["compare_log", str(log2_path.resolve())])
    ws.append(
        [
            "color_rule",
            "abs(pct_diff)<=1% green; 1%-5% yellow; >5% red; gray means missing/baseline zero",
        ]
    )
    headers = [
        "step",
        "focus_metric",
        f"{label1}_metric_key",
        f"{label2}_metric_key",
        f"{label1}_value",
        f"{label2}_value",
        f"{label2}_minus_{label1}",
        "pct_diff_vs_baseline",
        "status",
    ]
    ws.append(headers)

    for cell in ws[4]:
        apply_header_style(cell, COMPARE_FILL)

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=5, min_col=8, max_col=8):
        cell = row[0]
        cell.number_format = PCT_FORMAT
        status = ws.cell(row=cell.row, column=9).value
        fill_pct_cell(cell, cell.value, status)

    for row in ws.iter_rows(min_row=5, min_col=9, max_col=9):
        cell = row[0]
        if cell.value != "ok":
            cell.fill = MISSING_FILL

    ws.freeze_panes = "B5"
    ws.auto_filter.ref = ws.dimensions
    auto_fit_columns(ws)


def build_compare_matrix_rows(
    log1_steps: Dict[int, Dict[str, float | int]],
    log2_steps: Dict[int, Dict[str, float | int]],
) -> tuple[List[str], List[List[object]], List[int], List[List[str]]]:
    headers = ["step"]
    pct_diff_columns: List[int] = []
    status_rows: List[List[str]] = []

    for display_name, _ in SUMMARY_METRICS:
        headers.extend(
            [
                f"{display_name} | log1",
                f"{display_name} | log2",
                f"{display_name} | pct_diff",
            ]
        )
        pct_diff_columns.append(len(headers))

    rows: List[List[object]] = []
    all_steps = sorted(set(log1_steps) | set(log2_steps))

    for step in all_steps:
        metrics1 = log1_steps.get(step, {})
        metrics2 = log2_steps.get(step, {})
        row: List[object] = [step]
        status_row: List[str] = []

        for _, aliases in SUMMARY_METRICS:
            _, value1 = resolve_metric(metrics1, aliases)
            _, value2 = resolve_metric(metrics2, aliases)
            _, pct_diff, status = compare_values(value1, value2)
            row.extend([value1, value2, pct_diff])
            status_row.append(status)

        rows.append(row)
        status_rows.append(status_row)

    return headers, rows, pct_diff_columns, status_rows


def add_compare_chart(
    ws,
    *,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    pct_diff_columns: Sequence[int],
) -> None:
    chart = LineChart()
    chart.title = "Percent Diff Trend by Step"
    chart.x_axis.title = "Step"
    chart.y_axis.title = "Pct diff vs baseline"
    chart.y_axis.number_format = PCT_FORMAT
    chart.y_axis.scaling.min = -0.15
    chart.y_axis.scaling.max = 0.15
    chart.height = 12
    chart.width = 24
    chart.style = 2

    for col in pct_diff_columns:
        has_value = any(
            ws.cell(row=row_idx, column=col).value is not None
            for row_idx in range(data_start_row, data_end_row + 1)
        )
        if not has_value:
            continue
        data = Reference(ws, min_col=col, max_col=col, min_row=header_row, max_row=data_end_row)
        chart.add_data(data, titles_from_data=True)

    if chart.series:
        categories = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        chart.set_categories(categories)
        ws.add_chart(chart, f"A{data_end_row + 3}")


def write_compare_matrix_sheet(
    ws,
    *,
    log1_path: Path,
    log2_path: Path,
    label1: str,
    label2: str,
    headers: List[str],
    rows: List[List[object]],
    pct_diff_columns: Sequence[int],
    status_rows: Sequence[Sequence[str]],
) -> None:
    ws.append(["baseline_log", str(log1_path.resolve())])
    ws.append(["compare_log", str(log2_path.resolve())])
    ws.append(
        [
            "color_rule",
            "abs(pct_diff)<=1% green; 1%-5% yellow; >5% red; gray means missing/baseline zero",
        ]
    )
    ws.append(["compare_columns", f"step + ({label1} value, {label2} value, pct_diff) for each key metric"])
    ws.append(headers)

    for cell in ws[5]:
        apply_header_style(cell, COMPARE_FILL)

    for row in rows:
        ws.append(row)

    data_start_row = 6
    for row_offset, row_index in enumerate(range(data_start_row, data_start_row + len(rows))):
        statuses = status_rows[row_offset]
        for metric_index, pct_col in enumerate(pct_diff_columns):
            cell = ws.cell(row=row_index, column=pct_col)
            cell.number_format = PCT_FORMAT
            fill_pct_cell(cell, cell.value, statuses[metric_index])

    ws.freeze_panes = "B6"
    ws.auto_filter.ref = ws.dimensions
    auto_fit_columns(ws)

    if rows:
        add_compare_chart(
            ws,
            header_row=5,
            data_start_row=data_start_row,
            data_end_row=data_start_row + len(rows) - 1,
            pct_diff_columns=pct_diff_columns,
        )


def write_compare_detail_sheet(
    ws,
    *,
    log1_path: Path,
    log2_path: Path,
    label1: str,
    label2: str,
    rows: List[List[object]],
) -> None:
    ws.append(["baseline_log", str(log1_path.resolve())])
    ws.append(["compare_log", str(log2_path.resolve())])
    headers = [
        "step",
        "metric",
        f"{label1}_value",
        f"{label2}_value",
        f"{label2}_minus_{label1}",
        "pct_diff_vs_baseline",
        "status",
    ]
    ws.append(headers)

    for cell in ws[3]:
        apply_header_style(cell, COMPARE_FILL)

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=4, min_col=6, max_col=6):
        cell = row[0]
        cell.number_format = PCT_FORMAT
        status = ws.cell(row=cell.row, column=7).value
        fill_pct_cell(cell, cell.value, status)

    for row in ws.iter_rows(min_row=4, min_col=7, max_col=7):
        cell = row[0]
        if cell.value != "ok":
            cell.fill = MISSING_FILL

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = ws.dimensions
    auto_fit_columns(ws)


def build_workbook(
    log1_path: Path,
    log2_path: Path,
    output_path: Path,
    label1: str,
    label2: str,
) -> Path:
    log1_steps = parse_log(log1_path)
    log2_steps = parse_log(log2_path)
    all_metric_keys = sorted(
        set(collect_metric_keys(log1_steps)) | set(collect_metric_keys(log2_steps))
    )

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    log1_sheet = workbook.create_sheet(safe_sheet_title(f"{label1}_metrics"))
    write_log_sheet(
        log1_sheet,
        step_map=log1_steps,
        metric_keys=all_metric_keys,
        source_path=log1_path,
    )

    log2_sheet = workbook.create_sheet(safe_sheet_title(f"{label2}_metrics"))
    write_log_sheet(
        log2_sheet,
        step_map=log2_steps,
        metric_keys=all_metric_keys,
        source_path=log2_path,
    )

    summary_sheet = workbook.create_sheet("summary")
    summary_rows = build_summary_rows(log1_steps, log2_steps)
    write_summary_sheet(
        summary_sheet,
        log1_path=log1_path,
        log2_path=log2_path,
        label1=label1,
        label2=label2,
        rows=summary_rows,
    )

    compare_sheet = workbook.create_sheet("compare")
    compare_headers, compare_matrix_rows, pct_diff_columns, compare_status_rows = (
        build_compare_matrix_rows(log1_steps, log2_steps)
    )
    write_compare_matrix_sheet(
        compare_sheet,
        log1_path=log1_path,
        log2_path=log2_path,
        label1=label1,
        label2=label2,
        headers=compare_headers,
        rows=compare_matrix_rows,
        pct_diff_columns=pct_diff_columns,
        status_rows=compare_status_rows,
    )

    compare_detail_sheet = workbook.create_sheet("compare_detail")
    compare_rows = build_compare_rows(log1_steps, log2_steps, all_metric_keys)
    write_compare_detail_sheet(
        compare_detail_sheet,
        log1_path=log1_path,
        log2_path=log2_path,
        label1=label1,
        label2=label2,
        rows=compare_rows,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def main() -> int:
    args = parse_args()

    label1 = args.label1 or args.log1.stem or "log1"
    label2 = args.label2 or args.log2.stem or "log2"
    label1, label2 = ensure_distinct_labels(label1, label2)

    try:
        output_path = build_workbook(
            log1_path=args.log1,
            log2_path=args.log2,
            output_path=args.output,
            label1=label1,
            label2=label2,
        )
    except Exception as exc:  # pragma: no cover - CLI error path
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print(f"Excel file generated: {output_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
