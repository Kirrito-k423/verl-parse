#!/usr/bin/env python3
"""Scan a directory of VERL training logs and dump them into a single Excel.

Compared to ``parse_verl_logs_to_excel.py`` (which takes exactly two log files
and produces a side-by-side compare sheet) this script walks a directory
tree, parses every ``.log`` file it finds, and produces one workbook with:

* ``summary`` -- one row per log. The first column is a unique id built from
  the file's relative path + filename. Performance-related metrics come first
  (throughput at the first step + the last 10 steps + average excluding step
  1, plus per-stage timings), followed by the parallelism and batch/sequence
  hyperparameters that drive those numbers.
* ``log_<unique_id>`` -- one sheet per log with the full step-by-step
  metric grid (same layout as the existing single-log ``*_metrics`` sheet).
* ``metadata`` -- input folder, label mapping, parse timestamp, and notes.

Usage::

    python3 parse_verl_logs_dir.py testFold/ -o testFold_summary.xlsx
"""

from __future__ import annotations

import argparse
import ast
import datetime as _dt
import re
import statistics
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# ANSI stripping & step metric extraction
# ---------------------------------------------------------------------------

# Real ESC sequences. test2.log actually emits the *literal* text "[36m...[0m"
# so the parser is robust to both forms.
ANSI_ESCAPE_RE = re.compile(r"\x1b\[[0-?]*[ -/]*[@-~]")

# Matches both literal "[36m(pid=XXX)[0m " and the real escape-sequence form.
LOG_PREFIX_RE = re.compile(
    r"^(?:\x1b\[36m|\[36m)\s*\(?\s*[\w .+-]+pid=\d+\)?\s*(?:\x1b\[0m|\[0m)?\s*"
)
# Captures "key: value" pairs on a single line.
METRIC_RE = re.compile(
    r"(?P<key>[A-Za-z0-9_./@-]+):\s*"
    r"(?P<value>[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?)"
)


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="EAF4E4")
HYPER_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Per-step metric extraction
# ---------------------------------------------------------------------------


def _clean_line(raw_line: str) -> str:
    s = ANSI_ESCAPE_RE.sub("", raw_line).replace("\r", " ")
    return LOG_PREFIX_RE.sub("", s).strip()


def _to_number(raw_value: str) -> float | int:
    if re.fullmatch(r"[-+]?\d+", raw_value):
        return int(raw_value)
    return float(raw_value)


def parse_steps(log_path: Path) -> Dict[int, Dict[str, float | int]]:
    """Return ``{step: {metric: value}}`` for a single VERL log file."""

    if not log_path.is_file():
        raise FileNotFoundError(f"Log file not found: {log_path}")

    step_to_metrics: Dict[int, Dict[str, float | int]] = {}
    with log_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            line = _clean_line(raw_line)
            if "step:" not in line:
                continue
            metrics: Dict[str, float | int] = {}
            for match in METRIC_RE.finditer(line):
                metrics[match.group("key")] = _to_number(match.group("value"))

            step_value = metrics.get("step")
            if step_value is None or len(metrics) <= 1:
                continue
            step = int(step_value)
            if step in step_to_metrics:
                print(
                    f"Warning: duplicate step {step} in {log_path} at line {line_number}; "
                    "keeping the latest record.",
                    file=sys.stderr,
                )
            step_to_metrics[step] = metrics

    if not step_to_metrics:
        raise ValueError(f"No step metrics were found in {log_path}")
    return dict(sorted(step_to_metrics.items(), key=lambda kv: kv[0]))


# ---------------------------------------------------------------------------
# Config extraction from the printed Hydra dump
# ---------------------------------------------------------------------------


# Same prefix pattern as above, but anchored more loosely for the raw text.
_PLAIN_PREFIX_RE = re.compile(r"^\[36m\([^)]+pid=\d+\)\[0m\s*")
# A valid "key: value" or value-only line.
_VALUE_BODY = (
    r"(?:True|False|None|-?\d+(?:\.\d+)?(?:e[+-]?\d+)?|'[^']*'|\"[^\"]*\"|\{|\[)"
    r"(?:[\s,\}\]]*)$"
)
_KEY_VALUE_RE = re.compile(r"^[\s]*'[^']+'\s*:\s*" + _VALUE_BODY)
_BARE_VALUE_RE = re.compile(r"^[\s]*(?:True|False|None|-?\d+(?:\.\d+)?(?:e[+-]?\d+)?|'[^']*'|\"[^\"]*\"|\{|\}|\[|\])" + r"(?:[\s,\}\]]*)$")


def _strip_text_prefix(line: str) -> str:
    """Remove the literal "[36m(... pid=N)[0m " prefix (and any ANSI form)."""
    s = ANSI_ESCAPE_RE.sub("", line)
    s = _PLAIN_PREFIX_RE.sub("", s)
    return s


def _parse_hydra_dict(log_path: Path) -> Dict[str, Any]:
    """Best-effort parse of the printed config dict at the top of a VERL log.

    VERL dumps the active config once at startup as a multi-line ``str(dict)``.
    Some logs also have interleaved warnings that break up the literal.  We
    recover the dict by:

    1. taking every line that precedes the first ``step:N`` entry,
    2. dropping injected warning lines (file paths, UserWarning prose,
       and stray Python statements like ``use_critic=need_critic(config)``),
    3. joining the rest, and
    4. feeding the result to ``ast.literal_eval``.

    Returns an empty dict on any failure (so callers can fall back to a
    column of ``n/a`` rather than crashing the workbook build).
    """

    INJECTED_STMT_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_.]*\s*[=(]")

    raw_lines: List[str] = []
    step_idx: Optional[int] = None
    with log_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            raw_lines.append(line)
            if step_idx is None and re.search(r"step:\s*\d+", _clean_line(line)):
                step_idx = len(raw_lines) - 1
                break
    if step_idx is None or step_idx == 0:
        return {}

    cleaned: List[str] = []
    for raw in raw_lines[:step_idx]:
        s = _strip_text_prefix(raw).rstrip()
        if not s:
            continue
        cleaned.append(s)

    # Locate the dict boundaries.  Some logs have ``{'actor_rollout_ref'``
    # on the very first line of the dict; others put the opening brace on
    # the line before.
    start_idx: Optional[int] = None
    for i, line in enumerate(cleaned):
        if "{'actor_rollout_ref'" in line:
            start_idx = i
            break
    if start_idx is None:
        for i, line in enumerate(cleaned):
            if "actor_rollout_ref" in line and i > 0 and cleaned[i - 1].strip() == "{":
                start_idx = i - 1
                break
    if start_idx is None:
        return {}

    end_idx = len(cleaned)
    for i, line in enumerate(cleaned):
        if "[validate_config]" in line:
            end_idx = i
            break

    candidate_lines = cleaned[start_idx:end_idx]

    def _try_eval(lines: List[str]) -> Optional[Dict[str, Any]]:
        text = "\n".join(lines)
        try:
            parsed = ast.literal_eval(text)
        except (SyntaxError, ValueError):
            return None
        return parsed if isinstance(parsed, dict) else None

    direct = _try_eval(candidate_lines)
    if direct is not None:
        return direct

    sanitized: List[str] = []
    for line in candidate_lines:
        if "[36m(" in line and "pid=" in line:
            cut = line.find("[36m(")
            head = line[:cut]
            m = re.search(r"[,\}\]]\s*$", head)
            line = head[: m.end()] if m else ""
        s = line.strip()
        if not s:
            continue
        if INJECTED_STMT_RE.match(s):
            continue
        sanitized.append(line)

    cleaned_dict = _try_eval(sanitized)
    return cleaned_dict or {}


# Mapping of config keys we want to surface. Each entry: (label, dotted_path).
CONFIG_KEYS: Sequence[Tuple[str, str]] = (
    # --- actor / megatron parallelism ---
    ("actor_tp", "actor_rollout_ref.actor.megatron.tensor_model_parallel_size"),
    ("actor_pp", "actor_rollout_ref.actor.megatron.pipeline_model_parallel_size"),
    ("actor_cp", "actor_rollout_ref.actor.megatron.context_parallel_size"),
    ("actor_ep", "actor_rollout_ref.actor.megatron.expert_model_parallel_size"),
    ("actor_etp", "actor_rollout_ref.actor.megatron.expert_tensor_parallel_size"),
    ("actor_sp", "actor_rollout_ref.actor.megatron.sequence_parallel"),
    ("actor_use_remove_padding", "actor_rollout_ref.actor.megatron.use_remove_padding"),
    ("actor_grad_offload", "actor_rollout_ref.actor.megatron.grad_offload"),
    ("actor_optim_offload", "actor_rollout_ref.actor.megatron.optimizer_offload"),
    ("actor_ppo_epochs", "actor_rollout_ref.actor.ppo_epochs"),
    ("actor_ppo_mini_batch_size", "actor_rollout_ref.actor.ppo_mini_batch_size"),
    ("actor_ppo_micro_batch_size_per_gpu", "actor_rollout_ref.actor.ppo_micro_batch_size_per_gpu"),
    ("actor_ppo_max_token_len_per_gpu", "actor_rollout_ref.actor.ppo_max_token_len_per_gpu"),
    ("actor_clip_ratio", "actor_rollout_ref.actor.clip_ratio"),
    ("actor_kl_loss_coef", "actor_rollout_ref.actor.kl_loss_coef"),
    ("actor_entropy_coeff", "actor_rollout_ref.actor.entropy_coeff"),
    ("actor_lr", "actor_rollout_ref.actor.optim.lr"),
    # --- rollout ---
    ("rollout_tp", "actor_rollout_ref.rollout.tensor_model_parallel_size"),
    ("rollout_dp", "actor_rollout_ref.rollout.data_parallel_size"),
    ("rollout_expert_parallel_size", "actor_rollout_ref.rollout.expert_parallel_size"),
    ("rollout_n", "actor_rollout_ref.rollout.rollout_n"),
    ("rollout_prompt_length", "actor_rollout_ref.rollout.prompt_length"),
    ("rollout_response_length", "actor_rollout_ref.rollout.response_length"),
    ("rollout_max_num_batched_tokens", "actor_rollout_ref.rollout.max_num_batched_tokens"),
    ("rollout_max_num_seqs", "actor_rollout_ref.rollout.max_num_seqs"),
    ("rollout_gpu_memory_utilization", "actor_rollout_ref.rollout.gpu_memory_utilization"),
    ("rollout_enable_chunked_prefill", "actor_rollout_ref.rollout.enable_chunked_prefill"),
    ("rollout_enable_prefix_caching", "actor_rollout_ref.rollout.enable_prefix_caching"),
    ("rollout_enforce_eager", "actor_rollout_ref.rollout.enforce_eager"),
    ("rollout_use_torch_compile", "actor_rollout_ref.rollout.use_torch_compile"),
    ("rollout_dtype", "actor_rollout_ref.rollout.dtype"),
    # --- data / algorithm ---
    ("train_batch_size", "data.train_batch_size"),
    ("algorithm_adv_estimator", "algorithm.adv_estimator"),
    ("algorithm_use_kl_in_reward", "algorithm.use_kl_in_reward"),
    ("model_engine", "model_engine"),
)


def _dig(node: Any, dotted: str) -> Any:
    cur: Any = node
    for part in dotted.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return None
        cur = cur[part]
    return cur


def extract_config(parsed: Dict[str, Any]) -> Dict[str, Any]:
    return {label: _dig(parsed, path) for label, path in CONFIG_KEYS}


# ---------------------------------------------------------------------------
# Aggregation helpers
# ---------------------------------------------------------------------------


def _mean(values: Iterable[float]) -> Optional[float]:
    seq = list(values)
    if not seq:
        return None
    return statistics.fmean(seq)


def _step_throughput(steps: Dict[int, Dict[str, float | int]], step: int) -> Optional[float]:
    metrics = steps.get(step)
    if not metrics:
        return None
    val = metrics.get("perf/throughput")
    return float(val) if val is not None else None


def _avg_metric(steps: Dict[int, Dict[str, float | int]], key: str) -> Optional[float]:
    values = [float(m[key]) for m in steps.values() if key in m]
    return _mean(values)


def _avg_metric_excl_first(steps: Dict[int, Dict[str, float | int]], key: str) -> Optional[float]:
    if not steps:
        return None
    sorted_steps = sorted(steps.items())
    if len(sorted_steps) <= 1:
        return None
    values = [float(m[key]) for _, m in sorted_steps[1:] if key in m]
    return _mean(values)


def _last_n_steps(steps: Dict[int, Dict[str, float | int]], n: int) -> List[int]:
    keys = sorted(steps.keys())
    return keys[-n:] if len(keys) >= n else keys


# ---------------------------------------------------------------------------
# Excel building
# ---------------------------------------------------------------------------


def _safe_sheet_title(title: str, used: set) -> str:
    """Return a unique, Excel-safe sheet name (max 31 chars)."""
    # Excel forbids these characters in sheet names.
    cleaned = re.sub(r"[:\\/\?\*\[\]]", "_", title)
    # Drop file extension noise to leave room for uniqueness suffixes.
    for ext in (".log", ".txt", ".out", ".json"):
        if cleaned.lower().endswith(ext):
            cleaned = cleaned[: -len(ext)]
    cleaned = cleaned.strip("._-") or "sheet"
    base = cleaned[:31]
    candidate = base
    suffix = 2
    while candidate in used or not candidate:
        trimmed = base[: 31 - len(str(suffix)) - 1]
        candidate = f"{trimmed}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def _autofit_columns(ws, min_width: int = 10, max_width: int = 50) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            if cell.number_format and cell.number_format != "General":
                # numeric cells: width based on the format
                longest = max(longest, len(cell.number_format) + 2)
            else:
                longest = max(longest, min(len(text), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_width, min(longest + 2, max_width))


def _header_cell(cell, fill: PatternFill = HEADER_FILL) -> None:
    cell.font = BOLD
    cell.fill = fill
    cell.alignment = WRAP


def _safe_value(v: Any) -> Any:
    """Render config values nicely in Excel."""
    if v is None:
        return "n/a"
    if isinstance(v, bool):
        return "True" if v else "False"
    if isinstance(v, float):
        # Render floats-as-ints (e.g. 1.0/0.0) as ints when integral, but
        # keep the full precision for tiny learning rates like 5e-07.
        if v.is_integer():
            return int(v)
        if abs(v) >= 1e-3:
            return round(v, 6)
        return v
    return v


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------


# Order matters -- performance first, then hyperparameters.
TIMING_METRICS: Sequence[Tuple[str, str]] = (
    ("throughput", "perf/throughput"),
    ("time_per_step", "perf/time_per_step"),
    ("total_num_tokens", "perf/total_num_tokens"),
    ("mfu_actor", "perf/mfu/actor"),
    ("mfu_actor_infer", "perf/mfu/actor_infer"),
    ("timing_s/step", "timing_s/step"),
    ("timing_s/gen", "timing_s/gen"),
    ("timing_s/old_log_prob", "timing_s/old_log_prob"),
    ("timing_s/ref", "timing_s/ref"),
    ("timing_s/update_actor", "timing_s/update_actor"),
    ("timing_s/update_weights", "timing_s/update_weights"),
    ("timing_s/adv", "timing_s/adv"),
    ("timing_s/reward", "timing_s/reward"),
)


def build_summary_rows(
    log_records: List[Tuple[Path, Dict[int, Dict[str, float | int]], Dict[str, Any]]],
    root: Path,
) -> Tuple[List[str], List[List[Any]]]:
    """Return (column_headers, data_rows) for the summary sheet.

    The column layout is, in order:

    1. identity -- ``unique_id``, ``label``, ``rel_path``, ``n_steps``,
       ``first_step``, ``last_step``
    2. per-step throughputs -- step 1 plus the most recent 10 steps
       (``step1_throughput``, ``step{N-1}_throughput`` ... ``step{N-10}_throughput``)
       plus the throughput average excluding the first step
    3. per-stage timings -- ``avg_*_excl_step1`` columns for each of the
       stages the user cares about (gen, ref, old_log_prob, update_actor,
       update_weights, step, mfu, etc.)
    4. first-step timings -- the same per-stage values measured at step 1
       (warmup reference)
    5. hyperparameters -- actor / rollout / data parallelism and the
       sequence / batch / optimization knobs that drive performance
    """

    last_n = 10
    headers: List[str] = [
        "unique_id",
        "label",
        "rel_path",
        "n_steps",
        "first_step",
        "last_step",
        "step1_throughput",
    ]
    # Last-10 step throughputs, named by their actual step number, most
    # recent first.  Falls back to ``step{N-K}_throughput`` for K=1..10.
    for k in range(1, last_n + 1):
        headers.append(f"stepN_minus_{k}_throughput")
    for label, _ in TIMING_METRICS:
        headers.append(f"avg_{label}_excl_step1")
    for label, _ in TIMING_METRICS:
        headers.append(f"first_step_{label}")
    for label, _ in CONFIG_KEYS:
        headers.append(label)

    rows: List[List[Any]] = []
    for log_path, step_map, cfg in log_records:
        rel = log_path.relative_to(root)
        unique_id = str(rel)
        label = str(rel.parent) if rel.parent != Path(".") else rel.stem
        sorted_steps = sorted(step_map.keys())
        n_steps = len(sorted_steps)
        first_step = sorted_steps[0] if sorted_steps else None
        last_step = sorted_steps[-1] if sorted_steps else None

        row: List[Any] = [
            unique_id,
            label,
            str(rel),
            n_steps,
            first_step,
            last_step,
            _step_throughput(step_map, first_step) if first_step is not None else None,
        ]
        # Last-10 throughputs: most recent first.
        tail = _last_n_steps(step_map, last_n)
        for k in range(1, last_n + 1):
            if k <= len(tail):
                row.append(_step_throughput(step_map, tail[-k]))
            else:
                row.append(None)
        # Average throughput + per-stage timings (excluding first step).
        for _, key in TIMING_METRICS:
            row.append(_avg_metric_excl_first(step_map, key))
        # First-step timings as a warmup reference.
        first_metrics = step_map.get(first_step) if first_step is not None else None
        for _, key in TIMING_METRICS:
            if first_metrics and key in first_metrics:
                row.append(float(first_metrics[key]))
            else:
                row.append(None)
        # Config knobs.
        for label_, _ in CONFIG_KEYS:
            row.append(_safe_value(cfg.get(label_)))
        rows.append(row)
    return headers, rows


def write_summary_sheet(ws, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    ws.append(list(headers))
    for cell in ws[1]:
        _header_cell(cell, SUMMARY_FILL)

    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    # Apply number formats to throughput / timing columns.
    n_rows = len(rows)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        if "throughput" in header or "mfu" in header:
            ws.column_dimensions[col_letter].number_format = "0.00"
        elif header.startswith("first_") or header.startswith("avg_") or header == "first_step" or header == "last_step" or header == "n_steps":
            ws.column_dimensions[col_letter].number_format = "0.00"
        else:
            ws.column_dimensions[col_letter].number_format = "General"
        # Apply format to data rows as well so the cells display the right way.
        for row_idx in range(2, n_rows + 2):
            ws.cell(row=row_idx, column=col_idx).number_format = ws.column_dimensions[col_letter].number_format

    # Widen hyperparameter columns a bit.
    for col_idx, header in enumerate(headers, start=1):
        if any(header == lbl for lbl, _ in CONFIG_KEYS):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Per-log full sheet
# ---------------------------------------------------------------------------


def collect_metric_keys(*step_maps: Dict[int, Dict[str, float | int]]) -> List[str]:
    keys: set = set()
    for step_map in step_maps:
        for metrics in step_map.values():
            keys.update(metrics.keys())
    return sorted(keys)


def write_log_sheet(
    ws,
    *,
    label: str,
    step_map: Dict[int, Dict[str, float | int]],
    metric_keys: Sequence[str],
) -> None:
    ws.append(["step", *metric_keys])
    for cell in ws[1]:
        _header_cell(cell)

    for step in sorted(step_map.keys()):
        metrics = step_map[step]
        row = [step] + [metrics.get(k) for k in metric_keys]
        ws.append(row)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Metadata sheet
# ---------------------------------------------------------------------------


def write_metadata_sheet(
    ws,
    *,
    log_dir: Path,
    output_path: Path,
    records: Sequence[Tuple[Path, int]],
) -> None:
    """Write the metadata sheet.

    ``records`` is a list of ``(absolute_path, n_steps)`` so the per-file
    table can show the step count next to each log.
    """
    ws.append(["field", "value"])
    for cell in ws[1]:
        _header_cell(cell)
    rows: List[Tuple[str, str]] = [
        ("input_directory", str(log_dir.resolve())),
        ("output_xlsx", str(output_path.resolve())),
        ("parsed_at_utc", _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("log_count", str(len(records))),
    ]
    for k, v in rows:
        ws.append([k, v])

    # Per-file reference list (with relative id and step count).
    ws.append([])
    header_row_idx = ws.max_row + 1
    ws.append(["file_index", "absolute_path", "rel_id", "n_steps"])
    for cell in ws[header_row_idx]:
        _header_cell(cell, HYPER_FILL)
    for idx, (log_path, n_steps) in enumerate(records, start=1):
        try:
            rel_id = str(log_path.relative_to(log_dir))
        except ValueError:
            rel_id = log_path.name
        ws.append([idx, str(log_path.resolve()), rel_id, n_steps])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Top-level orchestration
# ---------------------------------------------------------------------------


def discover_logs(log_dir: Path) -> List[Path]:
    if not log_dir.exists():
        raise FileNotFoundError(f"Directory not found: {log_dir}")
    if not log_dir.is_dir():
        raise NotADirectoryError(f"Not a directory: {log_dir}")
    files = sorted(p for p in log_dir.rglob("*.log") if p.is_file())
    if not files:
        raise FileNotFoundError(f"No .log files under {log_dir}")
    return files


def build_workbook(log_dir: Path, output_path: Path) -> Path:
    files = discover_logs(log_dir)
    records: List[Tuple[Path, Dict[int, Dict[str, float | int]], Dict[str, Any]]] = []
    for log_path in files:
        try:
            step_map = parse_steps(log_path)
        except Exception as exc:
            print(f"  ! {log_path}: {exc}", file=sys.stderr)
            continue
        cfg = extract_config(_parse_hydra_dict(log_path))
        records.append((log_path, step_map, cfg))
        print(f"  parsed {log_path.relative_to(log_dir)} ({len(step_map)} steps)")

    if not records:
        raise RuntimeError(f"No logs could be parsed under {log_dir}")

    all_metric_keys = collect_metric_keys(*[step_map for _, step_map, _ in records])
    summary_headers, summary_rows = build_summary_rows(records, root=log_dir)

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_sheet = workbook.create_sheet("summary")
    write_summary_sheet(summary_sheet, summary_headers, summary_rows)

    used_sheet_titles: set = {"summary"}
    for log_path, step_map, _cfg in records:
        rel = log_path.relative_to(log_dir)
        sheet_id = str(rel).replace("/", "_").replace("\\", "_")
        title = _safe_sheet_title(f"log_{sheet_id}", used_sheet_titles)
        ws = workbook.create_sheet(title)
        write_log_sheet(ws, label=title, step_map=step_map, metric_keys=all_metric_keys)

    metadata_sheet = workbook.create_sheet("metadata")
    write_metadata_sheet(
        metadata_sheet,
        log_dir=log_dir,
        output_path=output_path,
        records=[(p, len(sm)) for p, sm, _ in records],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Walk a directory of VERL training logs and export one workbook with "
            "a summary sheet (first step + last 10 steps + averages + per-log config) "
            "and a per-log full-metrics sheet."
        )
    )
    parser.add_argument("logdir", type=Path, help="Directory containing .log files (recursive).")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("verl_log_summary.xlsx"),
        help="Output Excel path. Default: verl_log_summary.xlsx",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        out = build_workbook(args.logdir, args.output)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    print(f"\nExcel file generated: {out.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
